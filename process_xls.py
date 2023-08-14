from openpyxl import load_workbook
from process_geo import find_address, select_option


def add_city(city, filename):
    wb = load_workbook(filename)
    sheet = wb.active
    for row in range(1, sheet.max_row + 1):
        cell = f'A{row}'
        if not sheet[cell].value:
            continue
        if city + ',' not in sheet[cell].value:
            sheet[cell] = city + ', ' + sheet[cell].value
    wb.save(filename)


def fill_coordinates(filename):
    wb = load_workbook(filename)
    sheet = wb.active
    if 'coordinates' not in wb.sheetnames:
        new_sheet = wb.create_sheet('coordinates')
        not_found_addr = 0
        print('Поиск координат по адресу...')
        for row in range(1, sheet.max_row + 1):
            cell = f'A{row}'
            if not sheet[cell].value:
                continue
            print(f'Поиск: ' + sheet[cell].value)
            address_variants, geo_variants = find_address(sheet[cell].value)
            if address_variants:
                num = select_option(address_variants) if len(address_variants) > 1 else 0
                new_sheet[f'C{row}'] = geo_variants[num][0]
                new_sheet[f'B{row}'] = geo_variants[num][1]
                new_sheet[f'A{row}'] = address_variants[num]
            else:
                print('По адресу: ' + sheet[cell].value + ' данные не найдены!')
                not_found_addr += 1
        wb.save(filename)
        print(f'Общее число объектов: {sheet.max_row}\nЧисло объектов на выходе: {sheet.max_row - not_found_addr}')


def return_from_file(filename):
    address_list, geo_list = [], []
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name('coordinates')
    for cell in range(1, sheet.max_row + 1):
        x = sheet[f'C{cell}'].value
        y = sheet[f'B{cell}'].value
        address = sheet[f'A{cell}'].value
        if address:
            address_list.append(address)
        if x and y:
            geo_list.append((x, y))
    return address_list, geo_list


def handle_file(city_name, working_file):
    add_city(city_name, working_file)
    fill_coordinates(working_file)
    return return_from_file(working_file)
